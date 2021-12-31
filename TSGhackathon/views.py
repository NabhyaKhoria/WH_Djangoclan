from django.shortcuts import render,redirect
import math, random
from django.core.mail import EmailMessage
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.contrib import messages
from django.conf import settings
from django.contrib.auth import get_user_model, login, authenticate, logout
from event.models import *
from .models import *
import openpyxl
from datetime import date
from django.contrib.auth.models import User as defaultuser
from django.contrib.auth.models import auth

from django.core.paginator import PageNotAnInteger, Paginator

def home(request):
    return render(request, 'nav.html')


def base(request):
    
    studentWelfare = StudentWelfare.objects.all().order_by('date')
    technologys = Technology.objects.all().order_by('date')
    socials = Social.objects.all().order_by('date')
    sports = Sports.objects.all()

    events = []
    for stud in studentWelfare:
        events.append(stud)

    for stud in technologys:
        events.append(stud)

    for stud in socials:
        events.append(stud)

    for stud in sports:
        events.append(stud)
        
    now = date.today()
    date_new = []
    event2 = []
    for event in events:
        date_new.append(event.date)

    date_new.sort()

    for i in range(len(events)):
        for stu in events:
            if date_new[i] == stu.date:
                event2.append(stu)
                break
            else:
                pass

    student = []
    var_new = 0
    for event in event2:
        if event.date >= now:
            student.append(event)
            var_new += 1
        if var_new == 3:
            break
    if request.user.is_authenticated:
        print(request.user)
        P=Profile.objects.get(username=request.user)
        print("USer Successfullyy logged in ")
        print("USername= "+ str(request.user)+"has fullname ="+str(P.fullname) )
        print(P.fullname)
        context = {
        'student': student,
        'events': events,
        'studentWelfare': studentWelfare,
        'technologys':technologys,
        'socials':socials,
        'sports':sports,
        'user' :request.user,
        'profile':P,
        'event2':event2

        }
        return render(request, 'base.html', context)
    else:
        print(request.user)
        print("User Login Failed")
        context = {
        'student': student,
        'events': events,
        'studentWelfare': studentWelfare,
        'technologys':technologys,
        'socials':socials,
        'sports':sports,
        'user' :request.user,
        'event2':event2

        }
        return render(request, 'base.html', context)
    


def login1(request):
    if(request.method == 'POST'):
        OTP = ""
        email = request.POST['signin-email']
        wb_obj = openpyxl.load_workbook("media\student\Student Data.xlsx")
        sheet_obj = wb_obj.active
        semail=[]
        for j in range(3,sheet_obj.max_row+1):
            cell_obj = sheet_obj.cell(row = j, column = 1)
            semail.append(cell_obj.value)
        val=0
        print(semail)
        if semail.count(email):
            val=1
            string = '0123456789abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
            length = len(string)
            for i in range(6) :
                OTP += string[math.floor(random.random() * length)]
            subject = 'One Time Password (OTP) to login at TSG'
            email2 = EmailMessage(
                            subject, OTP, settings.EMAIL_HOST_USER, to=[email]
                        )
            email2.send(fail_silently=True)
            print(OTP)
        return render(request,'baselogin.html', {'email':email,'OTP':OTP,'val':val})
    else:
        return render(request,'base.html')


def loginuser(request):
    if(request.method == 'POST'):
        email = request.POST['signin-email']
        fname=""
        rollno=""
        OTP=request.POST['rotp']
        gotp=request.POST['otp']
        if OTP==gotp:
            wb_obj = openpyxl.load_workbook("media\student\Student Data.xlsx")
            sheet_obj = wb_obj.active
            for j in range(3,sheet_obj.max_row+1):
                cell_obj = sheet_obj.cell(row = j, column = 1)
                if cell_obj.value==email:
                    fname=sheet_obj.cell(row = j, column = 2).value
                    rollno=sheet_obj.cell(row = j, column = 3).value
            
            print("success")
            # print(User.objects.get(email=request.POST.get("signin-email")))
            # print(User.objects.get(email=email))
            try :
                a=User.objects.get(email=email)
                print(a)
                print("success101")
                print('success103')
                login(request, a)
                print('success105')

            except:
                print("success102")
                user=User.objects.create(email=email)
                user.save()
                login(request, user)
                profile=Profile.objects.create(username=user,fullname=fname,rollno=rollno )
                profile.save()


            if request.user.is_authenticated:
                print('SUCCESS AGAIN')
                return redirect('base')  
                
            
            
        else:
            print("fail")
    else:
        return render(request,'base.html')  

def student(request):
    achivement=Achivement.objects.all()
    t=Achivement.objects.filter(category='Technology')
    sc=Achivement.objects.filter(category='Social_Culture')
    sg=Achivement.objects.filter(category='Sports_Games')
    sw=Achivement.objects.filter(category='Social_Welfare')
    o=Achivement.objects.filter(category='Others')
    
    p=Paginator(achivement, 6) # creating paginator objects
    # getting the desired page number from url
    page_number = request.GET.get('page')
    try:
        page_obj = p.get_page(page_number) # returns desired page object
    except PageNotAnInteger:
        # if page_number is not an integer then assign the first page
        page_obj=p.page(1)
    except EmptyPage:
        # if page is empty then return last page
        page_obj = p.page(p.num_pages)
    P=Profile.objects.get(username=request.user)
    print("USer Successfullyy logged in ")
    print("USername= "+ str(request.user)+"has fullname ="+str(P.fullname) )
    print(P.fullname)
    context={
        'page_obj': page_obj,
        'Technology': t,
        'Social_Culture': sw,
        'Sports_Games': sg,
        'Student_Welfare': sw,
        'Others': o,
        'profile':P,
    }
    return render(request,'student.html',context)

def out(request):
    logout(request)
    return redirect('base')

def official_login(request):
    if(request.method == 'POST'):
        # email = request.POST['signin-email']
        username = request.POST.get('signup-username',False)
        password = request.POST.get('signup-password',False)
        wb_obj = openpyxl.load_workbook("media\official\Official.xlsx")
        sheet_obj = wb_obj.active
        semail=[]
        for j in range(2,sheet_obj.max_row+1):
            cell_obj = sheet_obj.cell(row = j, column = 3)
            semail.append(cell_obj.value)
        val=0
        print(semail)
        if semail.count(username):
            print("success")
            # print(User.objects.get(email=request.POST.get("signin-email")))
            # print(User.objects.get(email=email))
            # try :
                # a=User.objects.get(email=email)
                # a = defaultuser.objects.get(username=username,password=password)
            a = auth.authenticate(userename=username, password=password)
            print(a)
            print("success101")
            print('success103')
            # login(request, a)
            if a is not None:
                login(request,a)
            # print('success105')
            # except:
            # print("success102")
            else:
                user=User.objects.create_officialuser(email=username, password=password)
                user.save()
                print(user)
                login(request, user)
                profile=Profile.objects.create(username=user, fullname=username )
                profile.save()


            if request.user.is_authenticated:
                print('SUCCESS AGAIN')
                return redirect('base')
                
            
            
        else:
            print("fail")
            return render(request,'base.html')  
    else:
        # return render(request,'base.html')  
        pass